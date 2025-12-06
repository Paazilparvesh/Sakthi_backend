from django.shortcuts import render
from rest_framework.response import Response
from .models import (
    Admin,
    product_details,
    product_material,
    company,
    programer_details,
    Qa_details,
    acc_details,
    operator_details,
    material_type,
    Machine,
    All_User,
    Role,
)
from rest_framework import status
from .serializers import (
    product_detailsSerializer,
    product_materialSerializer,
    programer_detailsSerializer,
    Qa_detailsSerializer,
    material_typeSerializer,
    acc_detailsSerializer,
)
from rest_framework.decorators import api_view
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import (
    api_view,
    authentication_classes,
    permission_classes,
    permission_classes,
)
from django.db import transaction
from django.http import JsonResponse, HttpResponse
import json

import openpyxl
from openpyxl.utils import get_column_letter


# Create your views here.


# Login
@api_view(["POST"])
def login(request):
    username = request.data.get("username")
    password = request.data.get("password")

    try:
        user = All_User.objects.get(username=username)
    except All_User.DoesNotExist:
        return Response({"error": "User not found"}, status=404)

    if user.password != password:
        return Response({"error": "Invalid password"}, status=400)

    roles = user.role.values_list("name", flat=True)

    return Response(
        {
            "message": "Login successful",
            "username": user.username,
            "isAdmin": user.isAdmin,
            "roles": list(roles),
        }
    )


# Operators API'S
@api_view(["POST"])
def add_operator(request):
    try:
        data = request.data
        operator_name = data.get("operator_name")

        # ? Validation
        if not operator_name:
            return Response(
                {"status": False, "message": "Operator name is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        operator_obj = operator_details.objects.create(
            operator_name=operator_name,
        )
        return Response(
            {
                "msg": "Operator created successfully",
                "operator_name": operator_name,
            },
            status=status.HTTP_201_CREATED,
        )
    except Exception as e:
        print("? Error in add_operator:", str(e))
        return Response(
            {"status": False, "error": str(e)}, status=status.HTTP_400_BAD_REQUEST
        )


@api_view(["GET"])
def get_operator(request):
    try:
        operators = operator_details.objects.all()

        if not operators.exists():
            return Response(
                {"msg": "No operators found"}, status=status.HTTP_404_NOT_FOUND
            )

        # ? Serialize response
        result = []
        for op in operators:
            result.append(
                {
                    "id": op.id,
                    "operator_name": op.operator_name,
                }
            )

        return Response(result, status=status.HTTP_200_OK)

    except Exception as e:
        return Response(
            {"msg": f"Error fetching operators: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["PUT"])
def update_operator(request, operator_id):
    try:
        operator = operator_details.objects.get(id=operator_id)

        operator_name = request.data.get("operator_name")

        if not operator_name:
            return Response(
                {"status": False, "message": "Operator name is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        operator.operator_name = operator_name
        operator.save()

        return Response(
            {
                "status": True,
                "message": "Operator updated successfully",
                "operator_name": operator.operator_name,
            },
            status=status.HTTP_200_OK,
        )

    except operator_details.DoesNotExist:
        return Response(
            {"status": False, "message": "Operator not found"},
            status=status.HTTP_404_NOT_FOUND,
        )
    except Exception as e:
        return Response(
            {"status": False, "error": str(e)}, status=status.HTTP_400_BAD_REQUEST
        )


@api_view(["DELETE"])
def delete_operator(request, operator_id):
    try:
        operator = operator_details.objects.get(id=operator_id)
        operator.delete()

        return Response(
            {"status": True, "message": "Operator deleted successfully"},
            status=status.HTTP_200_OK,
        )

    except operator_details.DoesNotExist:
        return Response(
            {"status": False, "message": "Operator not found"},
            status=status.HTTP_404_NOT_FOUND,
        )
    except Exception as e:
        return Response(
            {"status": False, "error": str(e)}, status=status.HTTP_400_BAD_REQUEST
        )


# User
@api_view(["POST"])
def create_user(request):
    try:
        data = request.data
        username = data.get("username")
        email = data.get("email")
        password = data.get("password")
        isAdmin = data.get("isAdmin", False)
        roles_data = data.get("role", [])

        # ? Validation
        if not username:
            return Response(
                {"status": False, "message": "Username is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if All_User.objects.filter(username=username).exists():
            return Response(
                {"status": False, "message": "Username already exists."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user_obj = All_User.objects.create(
            username=username,
            email=email,
            password=password,
            isAdmin=isAdmin,
        )

        # Assign roles if provided
        for role_name in roles_data:
            role, created = Role.objects.get_or_create(name=role_name)
            user_obj.role.add(role)

        return Response(
            {
                "msg": "User created successfully",
                "username": username,
                "email": email,
                "isAdmin": isAdmin,
                "roles": roles_data,
            },
            status=status.HTTP_201_CREATED,
        )
    except Exception as e:
        print("? Error in create_user:", str(e))
        return Response(
            {"status": False, "error": str(e)}, status=status.HTTP_400_BAD_REQUEST
        )


@api_view(["GET"])
def get_all_users(request):
    try:
        all_users = All_User.objects.all()
        Response_data = {"total_users": all_users.count()}
        users_data = []
        for user in all_users:
            Response_data[str(user.id)] = [
                {
                    "username": user.username,
                    "email": user.email,
                    "roles": list(user.role.values_list("name", flat=True)),
                    "isAdmin": user.isAdmin,
                }
            ]

        return Response(Response_data, status=200)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(["PUT"])
def update_user(request, id):
    try:
        user = All_User.objects.get(id=id)

        username = request.data.get("username")
        email = request.data.get("email")
        password = request.data.get("password")
        isAdmin = request.data.get("isAdmin", user.isAdmin)
        roles_data = request.data.get("role", [])

        if not username:
            return Response(
                {"status": False, "message": "Username is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if All_User.objects.filter(username=username).exclude(id=id).exists():
            return Response(
                {"status": False, "message": "Username already exists."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user.username = username
        user.email = email
        if password:
            user.password = password
        user.isAdmin = isAdmin
        user.role.clear()

        for role_name in roles_data:
            role, created = Role.objects.get_or_create(name=role_name)
            user.role.add(role)

        user.save()

        return Response(
            {
                "status": True,
                "message": "User updated successfully",
                "username": user.username,
                "email": user.email,
                "isAdmin": user.isAdmin,
                "roles": roles_data,
            },
            status=status.HTTP_200_OK,
        )

    except All_User.DoesNotExist:
        return Response(
            {"status": False, "message": "User not found"},
            status=status.HTTP_404_NOT_FOUND,
        )
    except Exception as e:
        return Response(
            {"status": False, "error": str(e)}, status=status.HTTP_400_BAD_REQUEST
        )


@api_view(["DELETE"])
def delete_user(request, id):
    try:
        user = All_User.objects.get(id=id)
        user.delete()

        return Response(
            {"status": True, "message": "User deleted successfully"},
            status=status.HTTP_200_OK,
        )

    except All_User.DoesNotExist:
        return Response(
            {"status": False, "message": "User not found"},
            status=status.HTTP_404_NOT_FOUND,
        )
    except Exception as e:
        return Response(
            {"status": False, "error": str(e)}, status=status.HTTP_400_BAD_REQUEST
        )


# Companies API'S
@api_view(["GET"])
def get_companies(request):
    company_data = company.objects.all()
    get = []
    for comp in company_data:
        get.append(
            {
                "id": comp.id,
                "company_name": comp.company_name,
                "customer_name": comp.customer_name,
                "contact_no": comp.contact_no,
                "customer_dc_no": comp.customer_dc_no,
            }
        )
    return Response(get)


@api_view(["POST"])
def add_company(request):
    try:
        data = request.data
        company_name = data.get("company_name")
        customer_name = data.get("customer_name")
        contact_no = data.get("contact_no")
        customer_dc_no = data.get("customer_dc_no")

        # ? Validation
        if not company_name:
            return Response(
                {"status": False, "message": "Company name is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not customer_name:
            return Response(
                {"status": False, "message": "Customer name is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ? Create the company record
        new_company = company.objects.create(
            company_name=company_name,
            customer_name=customer_name,
            contact_no=contact_no,
            customer_dc_no=customer_dc_no,
        )

        return Response(
            {
                "status": True,
                "msg": "Company added successfully",
                "data": {
                    "id": new_company.id,
                    "company_name": new_company.company_name,
                    "customer_name": new_company.customer_name,
                    "contact_no": new_company.contact_no,
                    "customer_dc_no": new_company.customer_dc_no,
                },
            },
            status=status.HTTP_201_CREATED,
        )

    except Exception as e:
        return Response(
            {"status": False, "error": str(e)}, status=status.HTTP_400_BAD_REQUEST
        )


@api_view(["POST"])
def bulk_upload_company(request):
    companies_data = request.data.get("companies", [])

    if not companies_data:
        return Response({"msg": "Company list is required"}, status=400)

    created_count = 0
    updated_count = 0
    skipped_count = 0

    for data in companies_data:
        company_name = data.get("company_name")
        customer_name = data.get("customer_name")
        contact_no = data.get("contact_no", "")
        customer_dc_no = data.get("customer_dc_no", "")

        if not company_name or not customer_name:
            continue

        # ?? Look for existing
        existing = company.objects.filter(
            company_name__iexact=company_name.strip(),
            customer_name__iexact=customer_name.strip(),
        ).first()

        if existing:
            # ?? Check if all fields are identical ? skip
            if (
                existing.contact_no == contact_no
                and existing.customer_dc_no == customer_dc_no
            ):
                skipped_count += 1
                continue

            # ?? Something changed ? update fields
            existing.contact_no = contact_no
            existing.customer_dc_no = customer_dc_no
            existing.save()
            updated_count += 1

        else:
            # ? New company ? create
            company.objects.create(
                company_name=company_name,
                customer_name=customer_name,
                contact_no=contact_no,
                customer_dc_no=customer_dc_no,
            )
            created_count += 1

    return Response(
        {
            "msg": "Bulk upload completed",
            "created": created_count,
            "updated": updated_count,
            "skipped": skipped_count,
        }
    )


@api_view(["PUT"])
def update_company(request, pk):
    """
    Update a company by id (pk).
    Accepts JSON with fields to update:
    company_name, customer_name, company_address, contact_no, company_email, customer_dc_no
    """
    try:
        comp = company.objects.get(id=pk)
    except company.DoesNotExist:
        return Response(
            {"status": False, "message": "Company not found"},
            status=status.HTTP_404_NOT_FOUND,
        )

    data = request.data
    # Require company_name and customer_name to be non-empty
    company_name = data.get("company_name")
    customer_name = data.get("customer_name")

    if not company_name or not customer_name:
        return Response(
            {
                "status": False,
                "message": "company_name and customer_name are required.",
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        comp.company_name = company_name
        comp.customer_name = customer_name
        comp.contact_no = data.get("contact_no", comp.contact_no)
        comp.customer_dc_no = data.get("customer_dc_no", comp.customer_dc_no)
        comp.save()

        return Response(
            {
                "message": "Company updated successfully",
                "company": {
                    "id": comp.id,
                    "company_name": comp.company_name,
                    "customer_name": comp.customer_name,
                    "contact_no": comp.contact_no,
                    "customer_dc_no": comp.customer_dc_no,
                },
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        return Response(
            {"status": False, "error": str(e)}, status=status.HTTP_400_BAD_REQUEST
        )


@api_view(["DELETE"])
def delete_company(request, pk):
    """
    Delete a company by id (pk).
    """
    try:
        comp = company.objects.get(id=pk)
    except company.DoesNotExist:
        return Response(
            {"status": False, "message": "Company not found"},
            status=status.HTTP_404_NOT_FOUND,
        )

    try:
        comp.delete()
        return Response(
            {"status": True, "message": "Company deleted successfully"},
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        return Response(
            {"status": False, "error": str(e)}, status=status.HTTP_400_BAD_REQUEST
        )


# Material Type
@api_view(["GET"])
def get_material_type(request):
    types = material_type.objects.all().order_by("material_name")
    serializer = material_typeSerializer(types, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(["POST"])
def add_material_type(request):
    serializer = material_typeSerializer(data=request.data)

    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["PUT"])
def update_material_type(request, pk):
    try:
        instance = material_type.objects.get(id=pk)
    except material_type.DoesNotExist:
        return Response(
            {"error": "Material type not found"}, status=status.HTTP_404_NOT_FOUND
        )

    serializer = material_typeSerializer(instance, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["DELETE"])
def delete_material_type(request, pk):
    try:
        instance = material_type.objects.get(id=pk)
    except material_type.DoesNotExist:
        return Response(
            {"error": "Material type not found"}, status=status.HTTP_404_NOT_FOUND
        )

    instance.delete()
    return Response(
        {"message": "Material deleted successfully"}, status=status.HTTP_200_OK
    )


# Machines
@api_view(["GET"])
def get_machines(request):
    machines = Machine.objects.all().values("id", "machine")
    return Response(list(machines))


@api_view(["POST"])
def add_machine(request):
    machine_name = request.data.get("machine", "").strip()

    if not machine_name:
        return Response({"message": "Machine name is required"}, status=400)

    if Machine.objects.filter(machine__iexact=machine_name).exists():
        return Response({"message": "Machine already exists"}, status=409)

    Machine.objects.create(machine=machine_name)

    return Response({"message": "Machine added successfully"}, status=201)


@api_view(["PUT"])
def update_machine(request, id):
    machine_name = request.data.get("machine", "").strip()

    if not machine_name:
        return Response({"message": "Machine name is required"}, status=400)

    try:
        machine_obj = Machine.objects.get(id=id)
    except Machine.DoesNotExist:
        return Response({"message": "Machine not found"}, status=404)

    if Machine.objects.filter(machine__iexact=machine_name).exclude(id=id).exists():
        return Response({"message": "Machine already exists"}, status=409)

    machine_obj.machine = machine_name
    machine_obj.save()

    return Response({"message": "Machine updated successfully"}, status=200)


@api_view(["DELETE"])
def delete_machine(request, id):
    try:
        machine_obj = Machine.objects.get(id=id)
    except Machine.DoesNotExist:
        return Response({"message": "Machine not found"}, status=404)

    machine_obj.delete()

    return Response({"message": "Machine deleted successfully"}, status=200)


# Product API's
@api_view(["POST"])
def add_full_product(request):

    try:
        with transaction.atomic():

            # ? Step 1: Resolve inward user
            created_by_username = request.data.get("created_by")
            created_user = None

            if created_by_username:
                try:
                    created_user = All_User.objects.get(username=created_by_username)
                except All_User.DoesNotExist:
                    return Response(
                        {"msg": f"user '{created_by_username}' not found"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            # --- Product creation ---
            product = product_details.objects.create(
                serial_number=request.data.get("serial_number"),
                date=request.data.get("date"),
                inward_slip_number=request.data.get("inward_slip_number"),
                color=request.data.get("color"),
                worker_no=request.data.get("worker_no", "-"),
                company_name=request.data.get("company_name"),
                customer_name=request.data.get("customer_name"),
                customer_dc_no=request.data.get("customer_dc_no"),
                contact_no=request.data.get("contact_no"),
                created_by=created_user,
            )

            materials_data = request.data.get("materials", [])
            created_materials = []

            for mat in materials_data:

                # Validate numeric fields safely
                def safe_float(val):
                    try:
                        return float(val)
                    except (TypeError, ValueError):
                        return 0

                material = product_material.objects.create(
                    product_detail=product,
                    mat_type=mat.get("mat_type", ""),
                    mat_grade=mat.get("mat_grade", ""),
                    thick=safe_float(mat.get("thick")),
                    width=safe_float(mat.get("width")),
                    length=safe_float(mat.get("length")),
                    density=safe_float(mat.get("density")),
                    unit_weight=safe_float(mat.get("unit_weight")),
                    quantity=safe_float(mat.get("quantity")),
                    total_weight=safe_float(mat.get("total_weight")),
                    bay=mat.get("bay", ""),
                    stock_due=mat.get("stock_due", ""),
                    remarks=mat.get("remarks", "-"),
                )
                created_materials.append(material)

        # --- Serialize response ---
        response_data = {
            "msg": "Product and materials added successfully",
            "product": product_detailsSerializer(product).data,
            "materials": product_materialSerializer(created_materials, many=True).data,
        }
        return Response(response_data, status=status.HTTP_201_CREATED)

    except ValueError as ve:
        return Response({"msg": str(ve)}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response(
            {"msg": f"Error adding product: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
def get_full_products(request):
    """
    Fetch all products with their related materials.
    """
    try:
        products = product_details.objects.all().order_by("-id")

        if not products.exists():
            return Response(
                {"msg": "No products found"}, status=status.HTTP_404_NOT_FOUND
            )

        # ? Step 1: Build structured response
        result = []
        for prod in products:
            materials = product_material.objects.filter(product_detail=prod)

            result.append(
                {
                    "id": prod.id,
                    "serial_number": prod.serial_number,
                    "date": prod.date,
                    "inward_slip_number": prod.inward_slip_number,
                    "color": prod.color,
                    "worker_no": prod.worker_no,
                    "company_name": prod.company_name,
                    "customer_name": prod.customer_name,
                    "customer_dc_no": prod.customer_dc_no,
                    "contact_no": prod.contact_no,
                    "programer_status": prod.programer_status,
                    "outward_status": prod.outward_status,
                    "qa_status": prod.qa_status,
                    "created_by": prod.created_by.username if prod.created_by else None,
                    "materials": [
                        {
                            "id": mat.id,
                            "mat_type": mat.mat_type,
                            "mat_grade": mat.mat_grade,
                            "thick": mat.thick,
                            "width": mat.width,
                            "length": mat.length,
                            "density": mat.density,
                            "unit_weight": mat.unit_weight,
                            "quantity": mat.quantity,
                            "total_weight": mat.total_weight,
                            "bay": mat.bay,
                            "stock_due": mat.stock_due,
                            "remarks": mat.remarks,
                            "programer_status": mat.programer_status,
                            "qa_status": mat.qa_status,
                            "acc_status": mat.acc_status,
                        }
                        for mat in materials
                    ],
                }
            )

        return Response(result, status=status.HTTP_200_OK)

    except Exception as e:
        return Response(
            {"msg": f"Error fetching products: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


# Programmer API's
@api_view(["POST"])
def add_programer_Details(request):
    try:
        with transaction.atomic():

            data = request.data
            product_id = data.get("product_details")
            material_id = data.get("material_details")
            program_no = data.get("program_no")
            program_date = data.get("program_date")
            processed_quantity = data.get("processed_quantity")
            balance_quantity = data.get("balance_quantity")
            processed_width = data.get("processed_width")
            processed_length = data.get("processed_length")
            used_weight = data.get("used_weight")
            number_of_sheets = data.get("number_of_sheets")
            cut_length_per_sheet = data.get("cut_length_per_sheet")
            pierce_per_sheet = data.get("pierce_per_sheet")
            processed_mins_per_sheet = data.get("processed_mins_per_sheet")
            total_planned_hours = data.get("total_planned_hours")
            total_meters = data.get("total_meters")
            total_piercing = data.get("total_piercing")
            total_used_weight = data.get("total_used_weight")
            total_no_of_sheets = data.get("total_no_of_sheets")
            remarks = data.get("remarks")
            created_by = data.get("created_by")

            # ? Step 1: Resolve inward user
            if not created_by:
                return Response(
                    {"error": "Missing 'created_by' username."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            try:
                created_user = All_User.objects.get(username=created_by)
            except All_User.DoesNotExist:
                return Response(
                    {"error": f"user '{created_by}' not found"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            # ? Step 2: Validate product
            try:
                prod = product_details.objects.get(id=product_id)
            except product_details.DoesNotExist:
                return Response(
                    {"error": "Product details not found."},
                    status=status.HTTP_404_NOT_FOUND,
                )
            # ? Step 3: Validate material and link to product
            material_instance = None
            if material_id:
                try:
                    material_instance = product_material.objects.get(
                        id=material_id, product_detail=prod
                    )
                except product_material.DoesNotExist:
                    return Response(
                        {
                            "error": "Material not found or does not belong to this product."
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            programmer_obj = programer_details.objects.create(
                product_details=prod,
                material_details=material_instance,
                program_no=program_no,
                program_date=program_date,
                processed_quantity=processed_quantity,
                balance_quantity=balance_quantity,
                processed_width=processed_width,
                processed_length=processed_length,
                used_weight=used_weight,
                number_of_sheets=number_of_sheets,
                cut_length_per_sheet=cut_length_per_sheet,
                pierce_per_sheet=pierce_per_sheet,
                processed_mins_per_sheet=processed_mins_per_sheet,
                total_planned_hours=total_planned_hours,
                total_meters=total_meters,
                total_piercing=total_piercing,
                total_used_weight=total_used_weight,
                total_no_of_sheets=total_no_of_sheets,
                remarks=remarks,
                created_by=created_user,
            )

            # ? Step 7: Check if all materials are completed ? update product status
            all_materials = product_material.objects.filter(product_detail=prod)
            if all_materials.exists() and all(
                m.programer_status == "completed" for m in all_materials
            ):
                prod.programer_status = "completed"
            else:
                prod.programer_status = "pending"
            prod.save(update_fields=["programer_status"])

            return Response(
                {
                    "msg": "programmer created successfully",
                    "product_id": prod.id,
                    "material_id": material_instance.id if material_instance else None,
                    "program_no": program_no,
                    "program_date": program_date,
                    "processed_quantity": processed_quantity,
                    "balance_quantity": balance_quantity,
                    "processed_width": processed_width,
                    "processed_length": processed_length,
                    "used_weight": used_weight,
                    "number_of_sheets": number_of_sheets,
                    "cut_length_per_sheet": cut_length_per_sheet,
                    "pierce_per_sheet": pierce_per_sheet,
                    "processed_mins_per_sheet": processed_mins_per_sheet,
                    "total_planned_hours": total_planned_hours,
                    "total_meters": total_meters,
                    "total_piercing": total_piercing,
                    "total_used_weight": total_used_weight,
                    "total_no_of_sheets": total_no_of_sheets,
                    "remarks": remarks,
                    "created_by": created_user.username,
                    "programer_status": prod.programer_status,
                },
                status=status.HTTP_201_CREATED,
            )
    except Exception as e:
        print("? Error in add_account_new:", str(e))
        return Response(
            {"status": False, "error": str(e)}, status=status.HTTP_400_BAD_REQUEST
        )


@api_view(["GET"])
def get_programer_Details(request):
    """
    ? Fetch all programer_details, or filter by product_id / material_id
    Examples:
        GET /api/get_programer_Details/                          ? all records
        GET /api/get_programer_Details/?product_id=5             ? records for a specific product
        GET /api/get_programer_Details/?material_id=12           ? records for a specific material
        GET /api/get_programer_Details/?product_id=5&material_id=12 ? combined filter
    """
    try:
        product_id = request.query_params.get("product_id")
        material_id = request.query_params.get("material_id")

        # ? Base QuerySet
        details = programer_details.objects.all().select_related(
            "product_details", "material_details", "created_by"
        )

        # ? Apply filters dynamically
        if product_id:
            details = details.filter(product_details_id=product_id)
        if material_id:
            details = details.filter(material_details_id=material_id)

        if not details.exists():
            return Response(
                {"message": "No programmer details found for the given filters."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # ? Serialize and respond
        serializer = programer_detailsSerializer(details, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    except Exception as e:
        print("? Error in get_programer_Details:", str(e))
        return Response(
            {"error": str(e)},
            status=status.HTTP_400_BAD_REQUEST,
        )


@api_view(["POST"])
def create_pending_material(request):
    try:
        with transaction.atomic():
            data = request.data

            product_id = data.get("product_details")
            material_id = data.get("material_id")
            remaining_width = data.get("remaining_width")
            remaining_length = data.get("remaining_length")
            balance_qty = data.get("balance_quantity")  # ? NEW

            # STEP 1 � Validate product
            try:
                product = product_details.objects.get(id=product_id)
            except product_details.DoesNotExist:
                return Response(
                    {"error": "Product not found"}, status=status.HTTP_404_NOT_FOUND
                )

            # STEP 2 � Validate source material
            try:
                old_material = product_material.objects.get(
                    id=material_id, product_detail=product
                )
            except product_material.DoesNotExist:
                return Response(
                    {"error": "Material not found"}, status=status.HTTP_404_NOT_FOUND
                )

            # STEP 3 � Calculate new unit weight
            thick = old_material.thick or 0
            density = old_material.density or 0
            width = float(remaining_width)
            length = float(remaining_length)

            new_unit_weight = thick * width * length * density

            # STEP 4 � NEW QUANTITY + total weight
            qty = float(balance_qty) if balance_qty else 0  # ? replaced quantity
            total_weight = new_unit_weight * qty  # recalc

            # STEP 5 � Create new updated material
            new_mat = product_material.objects.create(
                product_detail=product,
                mat_type=old_material.mat_type,
                mat_grade=old_material.mat_grade,
                thick=old_material.thick,
                width=width,
                length=length,
                density=old_material.density,
                unit_weight=new_unit_weight,
                quantity=qty,  # ? USE NEW BALANCE QUANTITY
                total_weight=total_weight,  # recalculated
                bay=old_material.bay,
                stock_due=old_material.stock_due,
                remarks=old_material.remarks,
                status="pending",
                programer_status="pending",
                qa_status="pending",
                acc_status="pending",
            )

            return Response(
                {
                    "message": "Material copied and updated successfully",
                    "new_material_id": new_mat.id,
                    "width": width,
                    "length": length,
                    "quantity": qty,  # ? returned
                    "unit_weight": new_unit_weight,
                    "total_weight": total_weight,
                },
                status=status.HTTP_201_CREATED,
            )

    except Exception as e:
        print("? Error:", str(e))
        return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


# QA API's
@api_view(["POST"])
def add_qa_details(request):
    try:
        data = request.data
        processed_date = data.get("processed_date")
        shift = data.get("shift")
        no_of_sheets = data.get("no_of_sheets")
        cycletime_per_sheet = data.get("cycletime_per_sheet")
        total_cycle_time = data.get("total_cycle_time")
        machines_used = data.get("machines_used", [])
        product_id = data.get("product_details")
        material_id = data.get("material_details")
        created_by = data.get("created_by")

        # ? Get product_details instance
        try:
            prod = product_details.objects.get(id=product_id)
        except product_details.DoesNotExist:
            return Response(
                {"status": False, "message": "Product details not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        material_instance = None
        if material_id:
            try:
                material_instance = product_material.objects.get(
                    id=material_id, product_detail=prod
                )
            except product_material.DoesNotExist:
                return Response(
                    {
                        "status": False,
                        "message": "Material info not found or does not belong to this product.",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if Qa_details.objects.filter(
                material_details=material_instance, product_details=prod
            ).exists():
                return Response(
                    {
                        "status": False,
                        "message": "QA details already exist for this material and product.",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # ? Fetch creator instance
        created_by_obj = None

        try:
            created_by_obj = All_User.objects.get(username=created_by)
        except All_User.DoesNotExist:
            return Response(
                {"status": False, "message": "user not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # ? Create the QA details record
        qa_details_obj = Qa_details.objects.create(
            product_details=prod,
            material_details=material_instance,
            processed_date=processed_date,
            shift=shift,
            no_of_sheets=no_of_sheets,
            cycletime_per_sheet=cycletime_per_sheet,
            total_cycle_time=total_cycle_time,
            machines_used=machines_used,
            created_by=created_by_obj,
        )

        # ? Response
        return Response(
            {
                "msg": "QA details created successfully",
                "qa_id": qa_details_obj.id,
                "product_details": prod.id,
                "processed_date": processed_date,
                "shift": shift,
                "no_of_sheets": no_of_sheets,
                "cycletime_per_sheet": cycletime_per_sheet,
                "total_cycle_time": total_cycle_time,
                "machines_used": machines_used,
                "created_by": created_by_obj.username if created_by_obj else None,
            },
            status=status.HTTP_201_CREATED,
        )

    except Exception as e:
        return Response(
            {"status": False, "error": str(e)}, status=status.HTTP_400_BAD_REQUEST
        )


@api_view(["GET"])
def get_qa_details(request):
    """
    Returns serialized QA details with optional filters:
    - product_id
    - created_by_type ("QA" or "Accountant")
    - created_by_id
    """
    try:
        product_id = request.GET.get("product_id")
        creator_type = request.GET.get("created_by")

        # ? Build queryset
        qa_qs = Qa_details.objects.select_related(
            "product_details", "material_details", "created_by"
        ).all()

        # ? Apply filters
        if product_id:
            qa_qs = qa_qs.filter(product_details__id=product_id)

        # ? Serialize queryset
        serializer = Qa_detailsSerializer(qa_qs, many=True)

        # ? Return serialized data directly
        return Response(serializer.data, status=status.HTTP_200_OK)

    except Exception as e:
        print("? Error in get_qa_details:", str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# Accounts API's
@api_view(["POST"])
def add_acc_details(request):
    try:
        data = request.data
        invoice_no = data.get("invoice_no")
        status_acc = data.get("status")
        remarks = data.get("remarks")
        product_id = data.get("product_details")
        material_ids = data.get("material_details", [])  # <-- LIST
        created_by = data.get("created_by")

        # Validate product
        try:
            prod = product_details.objects.get(id=product_id)
        except product_details.DoesNotExist:
            return Response(
                {"status": False, "message": "Product not found"}, status=404
            )

        # Validate user
        try:
            created_by_obj = All_User.objects.get(username=created_by)
        except All_User.DoesNotExist:
            return Response({"status": False, "message": "User not found"}, status=404)

        created_records = []

        # LOOP THROUGH ALL MATERIALS
        for material_id in material_ids:
            try:
                material_instance = product_material.objects.get(
                    id=material_id, product_detail=prod
                )
            except product_material.DoesNotExist:
                return Response(
                    {
                        "status": False,
                        "message": f"Material {material_id} not found or not linked to product",
                    },
                    status=400,
                )

            acc_obj = acc_details.objects.create(
                product_details=prod,
                material_details=material_instance,
                invoice_no=invoice_no,
                status=status_acc,
                remarks=remarks,
                created_by=created_by_obj,
            )

            created_records.append(acc_obj.id)

        # Update outward status
        all_materials = product_material.objects.filter(product_detail=prod)
        if all(m.acc_status == "completed" for m in all_materials):
            prod.outward_status = "completed"
        else:
            prod.outward_status = "pending"
        prod.save()

        return Response(
            {
                "status": True,
                "msg": "Account details created successfully",
                "created_records": created_records,
            },
            status=201,
        )

    except Exception as e:
        return Response({"status": False, "error": str(e)}, status=400)


@api_view(["GET"])
def get_acc_details(request):
    try:
        product_id = request.GET.get("product_id")
        creator_by = request.GET.get("created_by")

        acc_qs = acc_details.objects.select_related(
            "product_details", "material_details", "created_by"
        ).all()

        # Filter by product id
        if product_id:
            acc_qs = acc_qs.filter(product_details__id=product_id)

        # Filter by created_by (username)
        if creator_by:
            acc_qs = acc_qs.filter(created_by__username=creator_by)

        serializer = acc_detailsSerializer(acc_qs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    except Exception as e:
        print("? Error in get_acc_details:", str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# Admin Page
@api_view(["GET"])
def get_role_list(request):
    try:
        roles = Role.objects.all()

        role_data = [
            {
                "id": role.id,
                "name": role.name,
            }
            for role in roles
        ]

        return Response(
            {
                "total_roles": roles.count(),
                "roles": role_data,
            },
            status=200,
        )
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(["GET"])
def total_product(request):
    count_product = product_details.objects.count()
    return Response({"Total Product": count_product}, status=200)


@api_view(["GET"])
def get_overall_details(request):
    """
    ? Fetches combined data from all models:
       - Users (all roles)
       - Companies
       - Products ? Materials ? Programer / QA / Accounts Details
    Optional filters:
    ?product_id=xx ? Fetch details for one product
    """
    try:
        product_id = request.query_params.get("product_id")
        products = product_details.objects.all().order_by("-id")
        if product_id:
            products = products.filter(id=product_id)

        product_data = []
        for prod in products:
            # Fetch materials
            materials = product_material.objects.filter(product_detail=prod)

            material_list = []
            for mat in materials:
                # Fetch Programmer, QA, and Accounts details linked to this material
                programer_info = list(
                    programer_details.objects.filter(material_details=mat).values(
                        "id",
                        "program_no",
                        "program_date",
                        "processed_quantity",
                        "balance_quantity",
                        "processed_width",
                        "processed_length",
                        "used_weight",
                        "number_of_sheets",
                        "cut_length_per_sheet",
                        "pierce_per_sheet",
                        "processed_mins_per_sheet",
                        "total_planned_hours",
                        "total_meters",
                        "total_piercing",
                        "total_used_weight",
                        "total_no_of_sheets",
                        "remarks",
                        "created_by__username",
                    )
                )

                qa_info = list(
                    Qa_details.objects.filter(material_details=mat).values(
                        "id",
                        "processed_date",
                        "shift",
                        "no_of_sheets",
                        "cycletime_per_sheet",
                        "total_cycle_time",
                        "operator_name",
                        "machines_used",
                        "created_by__username",
                    )
                )

                acc_info = list(
                    acc_details.objects.filter(material_details=mat).values(
                        "id", "invoice_no", "status", "remarks", "created_by__username"
                    )
                )

                material_list.append(
                    {
                        "id": mat.id,
                        "mat_type": mat.mat_type,
                        "mat_grade": mat.mat_grade,
                        "thick": mat.thick,
                        "width": mat.width,
                        "length": mat.length,
                        "density": mat.density,
                        "unit_weight": mat.unit_weight,
                        "quantity": mat.quantity,
                        "total_weight": mat.total_weight,
                        "bay": mat.bay,
                        "stock_due": mat.stock_due,
                        "remarks": mat.remarks,
                        "status": mat.status,
                        "programer_status": mat.programer_status,
                        "qa_status": mat.qa_status,
                        "acc_status": mat.acc_status,
                        "programer_details": programer_info,
                        "qa_details": qa_info,
                        "account_details": acc_info,
                    }
                )

            product_data.append(
                {
                    "product_id": prod.id,
                    "company_name": prod.company_name,
                    "serial_number": prod.serial_number,
                    "date": prod.date,
                    "inward_slip_number": prod.inward_slip_number,
                    "color": prod.color,
                    "worker_no": prod.worker_no,
                    "customer_name": prod.customer_name,
                    "customer_dc_no": prod.customer_dc_no,
                    "contact_no": prod.contact_no,
                    "programer_status": prod.programer_status,
                    "qa_status": prod.qa_status,
                    "outward_status": prod.outward_status,
                    "created_by": prod.created_by.username if prod.created_by else None,
                    "materials": material_list,
                }
            )
        response = product_data

        return Response(response, status=status.HTTP_200_OK)

    except Exception as e:
        print("? Error in get_overall_details:", str(e))
        return Response(
            {"status": False, "error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
def export_overall_details(request):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Overall Details"

        headers = [
            "Product ID",
            "Company",
            "Serial No",
            "Customer Name",
            "Material ID",
            "Mat Type",
            "Mat Grade",
            "Thickness",
            "Width",
            "Length",
            "Quantity",
            "Total Weight",
            "bay" "Programmer Status",
            "QA Status",
            "Account Status",
            "Program No",
            "Processed Qty",
            "Balance Qty",
            "Used Weight",
            "Sheets Count",
            "Program Created By",
            "QA Processed Date",
            "QA Shift",
            "QA Created By",
            "ACC Invoice No",
            "ACC Status",
            "ACC Created By",
        ]
        ws.append(headers)

        products = product_details.objects.all()

        for prod in products:
            materials = product_material.objects.filter(product_detail=prod)

            for mat in materials:
                prog = programer_details.objects.filter(material_details=mat).first()
                qa = Qa_details.objects.filter(material_details=mat).first()
                acc = acc_details.objects.filter(material_details=mat).first()

                ws.append(
                    [
                        prod.id,
                        prod.company_name,
                        prod.serial_number,
                        prod.customer_name,
                        mat.id,
                        mat.mat_type,
                        mat.mat_grade,
                        mat.thick,
                        mat.width,
                        mat.length,
                        mat.quantity,
                        mat.total_weight,
                        mat.bay,
                        mat.programer_status,
                        mat.qa_status,
                        mat.acc_status,
                        prog.program_no if prog else None,
                        prog.processed_quantity if prog else None,
                        prog.balance_quantity if prog else None,
                        prog.used_weight if prog else None,
                        prog.number_of_sheets if prog else None,
                        prog.created_by.username if prog and prog.created_by else None,
                        qa.processed_date if qa else None,
                        qa.shift if qa else None,
                        qa.created_by.username if qa and qa.created_by else None,
                        acc.invoice_no if acc else None,
                        acc.status if acc else None,
                        (
                            acc.created_by.username
                            if acc and acc.created_by.username
                            else None
                        ),
                    ]
                )

        for col in ws.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[get_column_letter(column)].width = max_length + 2

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = "overall_details_export.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    except Exception as e:
        print("? Export Error:", str(e))
        return Response({"error": str(e)}, status=500)


@api_view(["GET"])
def export_specific_details(request, ids=None):
    """
    Export Excel for all products OR specific products by path parameter.
    Example:
        /api/export_specific_details/[1,2,3]
        /api/export_specific_details/1,2,3?material_id=15
    """
    try:
        # 🌟 Get & clean material_id
        material_id = request.query_params.get("material_id")

        if material_id:
            cleaned = "".join(c for c in material_id if c.isdigit())
            if cleaned == "":
                return Response({"msg": "Invalid material_id"}, status=400)
            material_id = int(cleaned)

        # 🌟 Parse product IDs from path
        if ids:
            # Remove brackets or spaces: "[1,2,3]" → "1,2,3"
            cleaned = ids.replace("[", "").replace("]", "").replace(" ", "")

            # Split and convert to integers
            product_ids = [int(x) for x in cleaned.split(",") if x.isdigit()]

            if not product_ids:
                return Response({"msg": "Invalid product ID list"}, status=400)

            # Filter selected IDs
            products_qs = product_details.objects.filter(id__in=product_ids)

            if not products_qs.exists():
                return Response({"msg": "No products found for given IDs"}, status=404)

        else:
            # No IDs → export all
            products_qs = product_details.objects.all().order_by("-id")

        # -----------------------------
        # PREPARE EXCEL
        # -----------------------------
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Overall Details"

        headers = [
            "Product ID",
            "Inward Slip Number",
            "Company",
            "Serial No",
            "Customer Name",
            "Material ID",
            "Mat Type",
            "Mat Grade",
            "Thickness",
            "Width",
            "Length",
            "Quantity",
            "Total Weight",
            "Bay",
            "Programmer Status",
            "QA Status",
            "Account Status",
            "Program No",
            "Processed Qty",
            "Balance Qty",
            "Used Weight",
            "Sheets Count",
            "Program Created By",
            "QA Processed Date",
            "QA Shift",
            "QA Created By",
            "ACC Invoice No",
            "ACC Status",
            "ACC Created By",
        ]
        ws.append(headers)

        for prod in products_qs:
            materials = product_material.objects.filter(product_detail=prod)

            if material_id:
                materials = materials.filter(id=material_id)
                if not materials.exists():
                    return Response({"msg": "Material not found"}, status=404)

            for mat in materials:
                prog = programer_details.objects.filter(material_details=mat).first()
                qa = Qa_details.objects.filter(material_details=mat).first()
                acc = acc_details.objects.filter(material_details=mat).first()

                ws.append(
                    [
                        prod.id,
                        prod.inward_slip_number,
                        prod.company_name,
                        prod.serial_number,
                        prod.customer_name,
                        mat.id,
                        mat.mat_type,
                        mat.mat_grade,
                        mat.thick,
                        mat.width,
                        mat.length,
                        mat.quantity,
                        mat.total_weight,
                        mat.bay,
                        mat.programer_status,
                        mat.qa_status,
                        mat.acc_status,
                        prog.program_no if prog else None,
                        prog.processed_quantity if prog else None,
                        prog.balance_quantity if prog else None,
                        prog.used_weight if prog else None,
                        prog.number_of_sheets if prog else None,
                        prog.created_by.username if prog and prog.created_by else None,
                        qa.processed_date if qa else None,
                        qa.shift if qa else None,
                        qa.created_by.username if qa and qa.created_by else None,
                        acc.invoice_no if acc else None,
                        acc.status if acc else None,
                        acc.created_by.username if acc and acc.created_by else None,
                    ]
                )

        # Auto fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                txt = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(txt))
            ws.column_dimensions[col_letter].width = max_len + 2

        # Output file
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="overall_details_export.xlsx"'
        )

        wb.save(response)
        return response

    except Exception as e:
        print("❌ Export Error:", str(e))
        return Response({"error": str(e)}, status=500)


# update over details
@api_view(["PUT"])
def update_overall_details(request, product_id):
    try:
        # 1. Find product
        try:
            prod = product_details.objects.get(id=product_id)
        except product_details.DoesNotExist:
            return Response({"msg": "Product not found"}, status=404)

        # 2. Update Product Fields
        prod.company_name = request.data.get("company_name", prod.company_name)
        prod.serial_number = request.data.get("serial_number", prod.serial_number)
        prod.customer_name = request.data.get("customer_name", prod.customer_name)
        prod.contact_no = request.data.get("contact_no", prod.contact_no)
        prod.programer_status = request.data.get(
            "programer_status", prod.programer_status
        )
        prod.qa_status = request.data.get("qa_status", prod.qa_status)
        prod.outward_status = request.data.get("outward_status", prod.outward_status)
        prod.date = request.data.get("date", prod.date)
        prod.time = request.data.get("time", prod.time)
        prod.save()

        # 3. Update Materials (loop)
        materials_data = request.data.get("materials", [])

        for mat_data in materials_data:
            mat_id = mat_data.get("id")
            if not mat_id:
                continue

            try:
                mat = product_material.objects.get(id=mat_id, product_detail=prod)
            except product_material.DoesNotExist:
                continue

            mat.mat_type = mat_data.get("mat_type", mat.mat_type)
            mat.mat_grade = mat_data.get("mat_grade", mat.mat_grade)
            mat.thick = mat_data.get("thick", mat.thick)
            mat.width = mat_data.get("width", mat.width)
            mat.length = mat_data.get("length", mat.length)
            mat.quantity = mat_data.get("quantity", mat.quantity)
            mat.total_weight = mat_data.get("total_weight", mat.total_weight)
            mat.bay = mat_data.get("bay", mat.bay)
            mat.status = mat_data.get("status", mat.status)
            mat.programer_status = mat_data.get(
                "programer_status", mat.programer_status
            )
            mat.qa_status = mat_data.get("qa_status", mat.qa_status)
            mat.acc_status = mat_data.get("acc_status", mat.acc_status)
            mat.save()

            # --- Programmer Update ---
            program_data = mat_data.get("programer_details", [])
            for p in program_data:
                program_id = p.get("id")
                if program_id:
                    programer_details.objects.filter(id=program_id).update(
                        program_no=p.get("program_no"),
                        processed_quantity=p.get("processed_quantity"),
                        balance_quantity=p.get("balance_quantity"),
                        used_weight=p.get("used_weight"),
                        number_of_sheets=p.get("number_of_sheets"),
                        time=p.get("time"),
                        date=p.get("date"),
                        remarks=p.get("remarks"),
                    )

            # --- QA Update ---
            qa_data = mat_data.get("qa_details", [])
            for q in qa_data:
                qa_id = q.get("id")
                if qa_id:
                    Qa_details.objects.filter(id=qa_id).update(
                        processed_date=q.get("processed_date"),
                        time=q.get("time"),
                        shift=q.get("shift"),
                        operator_name=q.get("operator_name"),
                        remarks=q.get("remarks"),
                    )

            # --- Accounts Update ---
            acc_data = mat_data.get("account_details", [])
            for a in acc_data:
                acc_id = a.get("id")
                if acc_id:
                    acc_details.objects.filter(id=acc_id).update(
                        invoice_no=a.get("invoice_no"),
                        date=a.get("date"),
                        time=a.get("time"),
                        status=a.get("status"),
                        remarks=a.get("remarks"),
                    )

        return Response({"msg": "Updated successfully"}, status=200)

    except Exception as e:
        print("❌ Update Error:", str(e))
        return Response({"error": str(e)}, status=500)


# delete over all detail
@api_view(["DELETE"])
def delete_product(request, product_id):
    try:
        try:
            prod = product_details.objects.get(id=product_id)
        except product_details.DoesNotExist:
            return Response({"msg": "Product not found"}, status=404)

        prod.delete()

        return Response({"msg": "Product and all related data deleted"}, status=200)

    except Exception as e:
        print("❌ Delete Error:", str(e))
        return Response({"error": str(e)}, status=500)
